"""Excel export endpoints for UI pages (openpyxl)."""
from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal
from io import BytesIO
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.orm import Session

from app import crud, models, schemas
from app.database import get_db

router = APIRouter(prefix="/export", tags=["export"])

PART_TYPE_CN = {"standard": "标准件", "custom": "自制件", "assembly": "装配件"}
STATUS_CN = {"draft": "草稿", "released": "已发布", "obsolete": "已停用(历史)"}
TX_TYPE_CN = {"in": "入库", "out": "出库", "adjust": "调整"}

# 数量/金额单元格：保留 3 位小数（显示与写入一致）
DECIMAL_3_FMT = "0.000"


def _q3(x) -> float:
    if x is None:
        return 0.0
    d = Decimal(str(x)).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)
    return float(d)


def _apply_decimal3_range(ws, min_row: int, max_row: int, cols: tuple[int, ...]) -> None:
    for r in range(min_row, max_row + 1):
        for c in cols:
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is None or isinstance(v, str):
                continue
            if isinstance(v, (int, float, Decimal)):
                cell.number_format = DECIMAL_3_FMT


def _sheet_title(s: str, max_len: int = 31) -> str:
    bad = '[]:*?/\\'
    for c in bad:
        s = s.replace(c, "_")
    return s[:max_len] or "Sheet"


def _wb_response(wb: Workbook, filename_ascii: str, filename_cn: str | None = None) -> StreamingResponse:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{filename_ascii}"'}
    if filename_cn:
        headers["Content-Disposition"] += f"; filename*=UTF-8''{quote(filename_cn)}"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/dashboard")
def export_dashboard(db: Session = Depends(get_db)):
    from sqlalchemy import func as sqla_func

    total_materials = db.scalar(select(sqla_func.count(models.Material.id))) or 0
    standard_count = (
        db.scalar(
            select(sqla_func.count(models.Material.id)).where(models.Material.part_type == models.PartType.standard)
        )
        or 0
    )
    custom_count = (
        db.scalar(select(sqla_func.count(models.Material.id)).where(models.Material.part_type == models.PartType.custom)) or 0
    )
    assembly_count = (
        db.scalar(
            select(sqla_func.count(models.Material.id)).where(models.Material.part_type == models.PartType.assembly)
        )
        or 0
    )
    current_bom_count = (
        db.scalar(select(sqla_func.count(models.BOMHeader.id)).where(models.BOMHeader.is_current.is_(True))) or 0
    )
    low_stock_count = (
        db.scalar(
            select(sqla_func.count(models.Material.id)).where(models.Material.current_stock < models.Material.safety_stock)
        )
        or 0
    )
    current_revision_count = (
        db.scalar(select(sqla_func.count(models.PartRevision.id)).where(models.PartRevision.is_current.is_(True))) or 0
    )

    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_title("仪表盘")
    ws.append(["指标", "数值"])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    rows = [
        ("物料总数", total_materials),
        ("标准件", standard_count),
        ("自制件", custom_count),
        ("装配件", assembly_count),
        ("当前生效BOM数", current_bom_count),
        ("低库存物料数", low_stock_count),
        ("当前生效版本数", current_revision_count),
        ("导出时间", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")),
    ]
    for r in rows:
        ws.append(list(r))
    return _wb_response(wb, "dashboard.xlsx", "仪表盘.xlsx")


def _filter_materials(db: Session, view: str):
    q = select(models.Material).order_by(models.Material.id.desc())
    rows = list(db.scalars(q).all())
    if view == "standard":
        rows = [m for m in rows if m.part_type == models.PartType.standard]
    elif view == "nonstandard":
        rows = [m for m in rows if m.part_type in (models.PartType.custom, models.PartType.assembly)]
    return rows


@router.get("/materials")
def export_materials(view: str = Query(default="all", pattern="^(all|standard|nonstandard)$"), db: Session = Depends(get_db)):
    materials = _filter_materials(db, view)
    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_title("物料")
    headers = [
        "编码",
        "名称",
        "型号/参数",
        "物料类型",
        "分类",
        "封装",
        "库位",
        "单位",
        "当前库存",
        "安全库存",
        "单价",
        "默认供应商",
        "备注",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).font = Font(bold=True)
    for m in materials:
        spec_draw = " / ".join(x for x in [(m.spec or "").strip(), (m.drawing_no or "").strip()] if x)
        ws.append(
            [
                m.code,
                m.name,
                spec_draw,
                m.material_type or "",
                m.category or "",
                m.package_name or "",
                m.storage_location or "",
                m.unit or "",
                _q3(m.current_stock),
                _q3(m.safety_stock),
                _q3(m.unit_price),
                m.default_supplier or "",
                (m.remark or "")[:5000],
            ]
        )
    if ws.max_row >= 2:
        _apply_decimal3_range(ws, 2, ws.max_row, (8, 9, 10))
    name = f"物料_{view}.xlsx"
    return _wb_response(wb, f"materials_{view}.xlsx", name)


@router.get("/categories")
def export_categories(
    option_type: str = Query(
        default="category",
        alias="type",
        pattern="^(category|unit|material_type|material_attr|grade|product_category)$",
    ),
    db: Session = Depends(get_db),
):
    wb = Workbook()
    ws = wb.active
    if option_type == "category":
        ws.title = _sheet_title("分类")
        ws.append(["名称", "编码前缀", "排序", "启用", "备注"])
        for c in range(1, 6):
            ws.cell(1, c).font = Font(bold=True)
        for row in crud.list_material_categories(db):
            ws.append([row.name, row.code_prefix, row.sort_order, "是" if row.is_active else "否", row.remark or ""])
        fname_cn = "类别_分类.xlsx"
        fname_en = "categories_category.xlsx"
    else:
        ws.title = _sheet_title(option_type)
        ws.append(["名称", "排序", "启用", "备注"])
        for c in range(1, 5):
            ws.cell(1, c).font = Font(bold=True)
        for row in crud.list_system_options(db, option_type):
            ws.append([row.name, row.sort_order, "是" if row.is_active else "否", row.remark or ""])
        fname_cn = f"类别_{option_type}.xlsx"
        fname_en = f"categories_{option_type}.xlsx"
    return _wb_response(wb, fname_en, fname_cn)


@router.get("/revisions/{material_id}")
def export_revisions(material_id: int, db: Session = Depends(get_db)):
    rows = crud.list_revisions(db, material_id)
    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_title("版本")
    headers = [
        "版本号",
        "图号",
        "PDF路径",
        "模型路径",
        "状态",
        "当前版本",
        "用途",
        "材质",
        "标准",
        "等级",
        "变更说明",
        "创建时间",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).font = Font(bold=True)
    for r in rows:
        st = r.status.value if hasattr(r.status, "value") else str(r.status)
        ws.append(
            [
                r.revision,
                r.drawing_no or "",
                r.file_path_pdf or "",
                r.file_path_model or "",
                STATUS_CN.get(st, st),
                "是" if r.is_current else "否",
                r.purpose or "",
                r.material_name or "",
                r.standard or "",
                r.grade or "",
                (r.change_note or "")[:5000],
                r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else "",
            ]
        )
    return _wb_response(wb, f"revisions_{material_id}.xlsx", f"版本_物料{material_id}.xlsx")


@router.get("/boms")
def export_boms(db: Session = Depends(get_db)):
    boms = crud.list_boms(db)
    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_title("BOM列表")
    ws.append(["产品编码", "产品名称", "BOM版本", "版本说明", "状态", "当前生效", "创建时间", "更新时间"])
    for c in range(1, 9):
        ws.cell(1, c).font = Font(bold=True)
    for b in boms:
        st = b.status.value if hasattr(b.status, "value") else str(b.status)
        ws.append(
            [
                b.product_code,
                b.product_name,
                b.bom_version,
                b.revision_note or "",
                STATUS_CN.get(st, st),
                "是" if b.is_current else "否",
                b.created_at.strftime("%Y-%m-%d %H:%M:%S") if b.created_at else "",
                b.updated_at.strftime("%Y-%m-%d %H:%M:%S") if b.updated_at else "",
            ]
        )
    return _wb_response(wb, "boms.xlsx", "BOM列表.xlsx")


@router.get("/bom/{bom_id}")
def export_bom_detail(bom_id: int, db: Session = Depends(get_db)):
    detail = crud.get_bom_detail(db, bom_id)
    h = detail["header"]
    items = detail["items"]
    total = detail["total_cost"]
    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_title("BOM明细")
    st = h.status.value if hasattr(h.status, "value") else str(h.status)
    ws.append(["BOM主信息"])
    ws.append(["产品编码", h.product_code])
    ws.append(["产品名称", h.product_name])
    ws.append(["BOM版本", h.bom_version])
    ws.append(["版本说明", h.revision_note or ""])
    ws.append(["状态", STATUS_CN.get(st, st)])
    ws.append(["当前生效", "是" if h.is_current else "否"])
    total_cost_row = ws.max_row + 1
    ws.append(["总成本", _q3(total)])
    ws.append([])
    headers = [
        "行号",
        "物料编码",
        "物料名称",
        "规格/图号",
        "用途",
        "数量",
        "单价",
        "总价",
        "备注",
    ]
    ws.append(headers)
    row_start = ws.max_row
    for c in range(1, len(headers) + 1):
        ws.cell(row_start, c).font = Font(bold=True)
    detail_start = ws.max_row + 1
    for it in items:
        spec = getattr(it, "spec_drawing", None) or ""
        ws.append(
            [
                it.line_no,
                it.material_code,
                it.material_name,
                spec,
                it.usage or "",
                _q3(it.qty),
                _q3(it.unit_price),
                _q3(it.total_price),
                (it.remark or "")[:2000],
            ]
        )
    ws.cell(row=total_cost_row, column=2).number_format = DECIMAL_3_FMT
    if ws.max_row >= detail_start:
        _apply_decimal3_range(ws, detail_start, ws.max_row, (6, 7, 8))
    return _wb_response(wb, f"bom_{bom_id}.xlsx", f"BOM明细_{bom_id}.xlsx")


@router.get("/inventory")
def export_inventory(
    material_id: int | None = Query(default=None),
    transaction_type: str | None = Query(default=None),
    db: Session = Depends(get_db),
):
    if transaction_type and transaction_type not in ("in", "out", "adjust"):
        raise HTTPException(status_code=422, detail="transaction_type must be in, out, or adjust")
    if material_id is not None:
        txs = crud.list_transactions_by_material(db, material_id)
    else:
        txs = crud.list_transactions(db)
    if transaction_type:
        txs = [t for t in txs if (t.transaction_type.value if hasattr(t.transaction_type, "value") else str(t.transaction_type)) == transaction_type]

    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_title("库存流水")
    ws.append(["ID", "物料ID", "类型", "数量", "单价", "来源类型", "来源单号", "备注", "创建时间"])
    for c in range(1, 10):
        ws.cell(1, c).font = Font(bold=True)
    for t in txs:
        tt = t.transaction_type.value if hasattr(t.transaction_type, "value") else str(t.transaction_type)
        ws.append(
            [
                t.id,
                t.material_id,
                TX_TYPE_CN.get(tt, tt),
                _q3(t.qty),
                _q3(t.unit_price),
                t.reference_type or "",
                t.reference_no or "",
                (t.remark or "")[:2000],
                t.created_at.strftime("%Y-%m-%d %H:%M:%S") if t.created_at else "",
            ]
        )
    if ws.max_row >= 2:
        _apply_decimal3_range(ws, 2, ws.max_row, (4, 5))
    return _wb_response(wb, "inventory.xlsx", "库存流水.xlsx")


@router.get("/suppliers")
def export_suppliers(db: Session = Depends(get_db)):
    suppliers = crud.list_suppliers(db)
    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_title("供应商")
    ws.append(
        [
            "供应商编码",
            "公司名称",
            "供应商分类",
            "账期(天)",
            "统一信用代码",
            "开户行",
            "账号",
            "行号",
            "联系人",
            "电话",
            "地址",
            "状态",
        ]
    )
    for c in range(1, 13):
        ws.cell(1, c).font = Font(bold=True)
    for s in suppliers:
        cats = "、".join(s.get("supplier_categories", []))
        ptd = s.get("payment_term_days")
        ws.append(
            [
                s.get("supplier_code") or "",
                s.get("company_name") or "",
                cats,
                ptd if ptd is not None else "",
                s.get("credit_code") or "",
                s.get("bank_name") or "",
                s.get("bank_account") or "",
                s.get("bank_no") or "",
                s.get("contact_person") or "",
                s.get("phone") or "",
                s.get("address") or "",
                "启用" if s.get("is_active") else "停用",
            ]
        )
    return _wb_response(wb, "suppliers.xlsx", "供应商.xlsx")


@router.get("/procurement")
def export_procurement(
    bom_id: int = Query(..., ge=1),
    production_qty: Decimal = Query(..., gt=0),
    db: Session = Depends(get_db),
):
    result = crud.calc_shortage(db, schemas.ShortageCalcRequest(bom_id=bom_id, production_qty=production_qty))

    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_title("缺料明细")
    ws.append(
        ["BOM ID", bom_id, "计划产量", _q3(production_qty), "总采购金额", _q3(result.total_estimated_cost)]
    )
    _apply_decimal3_range(ws, 1, 1, (4, 6))
    ws.append([])
    headers = [
        "编码",
        "名称",
        "规格/图号",
        "供应商",
        "单套用量",
        "总需求量",
        "当前库存",
        "安全库存",
        "安全缺口",
        "清库缺口",
        "建议采购量",
        "单价",
        "预估金额",
    ]
    ws.append(headers)
    row_h = ws.max_row
    for c in range(1, len(headers) + 1):
        ws.cell(row_h, c).font = Font(bold=True)
    shortage_first = ws.max_row + 1
    for it in result.shortage_list:
        ws.append(
            [
                it.material_code,
                it.material_name,
                it.spec_drawing or "",
                it.default_supplier or "",
                _q3(it.unit_usage),
                _q3(it.total_required_qty),
                _q3(it.current_stock),
                _q3(it.safety_stock),
                _q3(it.safety_shortage_qty),
                _q3(it.clear_shortage_qty),
                _q3(it.suggested_purchase_qty),
                _q3(it.unit_price),
                _q3(it.estimated_amount),
            ]
        )
    if ws.max_row >= shortage_first:
        _apply_decimal3_range(ws, shortage_first, ws.max_row, (5, 6, 7, 8, 9, 10, 11, 12, 13, 14))

    ws2 = wb.create_sheet(_sheet_title("按供应商"))
    ws2.append(["供应商", "小计金额", "物料摘要"])
    for c in range(1, 4):
        ws2.cell(1, c).font = Font(bold=True)
    for g in result.grouped_by_supplier:
        summary = "，".join([f"{i.material_code}({_q3(i.suggested_purchase_qty):.3f})" for i in g.items[:50]])
        if len(g.items) > 50:
            summary += "..."
        ws2.append([g.supplier, _q3(g.supplier_total_amount), summary])

    if ws2.max_row >= 2:
        _apply_decimal3_range(ws2, 2, ws2.max_row, (2,))
    return _wb_response(wb, f"procurement_{bom_id}.xlsx", f"缺料采购_{bom_id}.xlsx")


@router.get("/inquiry/{inquiry_id}")
def export_inquiry(inquiry_id: int, db: Session = Depends(get_db)):
    inq = crud.get_inquiry(db, inquiry_id)
    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_title("询价单")

    ws.append(["询价单号", inq.inquiry_no, "状态", inq.status, "日期", str(inq.inquiry_date)[:10]])
    ws.append(["供应商", inq.supplier_company or "", "联系人", inq.supplier_contact or "", "电话", inq.supplier_phone or ""])
    ws.append(["有效期至", str(inq.valid_until)[:10] if inq.valid_until else "", "交货地址", inq.delivery_address or "", "付款条件", inq.payment_terms or ""])
    ws.append(["备注", inq.header_remark or ""])
    ws.append([])
    headers = [
        "行号",
        "物料编码",
        "名称",
        "规格/图号",
        "材质",
        "等级",
        "单位",
        "数量",
        "单价",
        "总价",
        "备注",
    ]
    ws.append(headers)
    rh = ws.max_row
    for c in range(1, len(headers) + 1):
        ws.cell(rh, c).font = Font(bold=True)
    start = ws.max_row + 1
    for ln in inq.lines:
        ws.append([
            ln.line_no,
            ln.material_code,
            ln.material_name,
            ln.spec_drawing or "",
            ln.material_name_attr or "",
            ln.grade_attr or "",
            ln.unit or "",
            _q3(ln.qty),
            "",
            "",
            ln.remark or "",
        ])
    if ws.max_row >= start:
        _apply_decimal3_range(ws, start, ws.max_row, (8,))
    ws.append([])
    ws.append(["合计金额（供方填报）", "—"])
    return _wb_response(wb, f"inquiry_{inq.inquiry_no}.xlsx", f"询价单_{inq.inquiry_no}.xlsx")
