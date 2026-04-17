"""Excel import endpoints (openpyxl). Column headers align with /export/materials."""
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from pydantic import ValidationError
from sqlalchemy.orm import Session

from app import crud, models, schemas
from app.database import get_db

router = APIRouter(prefix="/import", tags=["import"])

MAX_IMPORT_BYTES = 5 * 1024 * 1024

# 与导出列一致；额外可选列便于补全主数据
MATERIAL_IMPORT_HEADERS = [
    "编码",
    "名称",
    "型号/参数",
    "物料类型",
    "分类",
    "封装",
    "库位",
    "单位",
    "当前库存",
    "安全库存",
    "单价",
    "默认供应商",
    "备注",
]

MATERIAL_HEADER_TO_FIELD: dict[str, str] = {
    "编码": "code",
    "code": "code",
    "名称": "name",
    "name": "name",
    "型号/参数": "spec",
    "物料类型": "material_type",
    "material_type": "material_type",
    "规格": "spec",
    "spec": "spec",
    "规格/图号": "spec_drawing",
    "spec_drawing": "spec_drawing",
    "单位": "unit",
    "unit": "unit",
    "分类": "category",
    "category": "category",
    "封装": "package_name",
    "package_name": "package_name",
    "库位": "storage_location",
    "storage_location": "storage_location",
    "类型": "part_type",
    "part_type": "part_type",
    "用途": "usage",
    "usage": "usage",
    "材质": "material_name_attr",
    "material_name_attr": "material_name_attr",
    "等级": "grade_attr",
    "grade_attr": "grade_attr",
    "默认供应商": "default_supplier",
    "default_supplier": "default_supplier",
    "采购链接": "purchase_link",
    "purchase_link": "purchase_link",
    "当前版本": "current_revision",
    "current_revision": "current_revision",
    "编辑状态": "status",
    "status": "status",
    "启用状态": "is_active",
    "is_active": "is_active",
    "当前库存": "current_stock",
    "current_stock": "current_stock",
    "安全库存": "safety_stock",
    "safety_stock": "safety_stock",
    "单价": "unit_price",
    "unit_price": "unit_price",
    "税率": "tax_rate",
    "tax_rate": "tax_rate",
    "备注": "remark",
    "remark": "remark",
    "标准": "standard_attr",
    "standard_attr": "standard_attr",
    "图号": "drawing_no",
    "drawing_no": "drawing_no",
}

PART_TYPE_IMPORT = {
    "标准件": models.PartType.standard,
    "自制件": models.PartType.custom,
    "装配件": models.PartType.assembly,
    "standard": models.PartType.standard,
    "custom": models.PartType.custom,
    "assembly": models.PartType.assembly,
}

STATUS_IMPORT = {
    "草稿": models.StatusType.draft,
    "已发布": models.StatusType.released,
    "已停用(历史)": models.StatusType.obsolete,
    "obsolete": models.StatusType.obsolete,
    "draft": models.StatusType.draft,
    "released": models.StatusType.released,
}


def _sheet_title(s: str, max_len: int = 31) -> str:
    bad = '[]:*?/\\'
    for c in bad:
        s = s.replace(c, "_")
    return s[:max_len] or "Sheet"


def _norm_header(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _cell_str(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return "是" if v else "否"
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    if isinstance(v, int) and not isinstance(v, bool):
        return str(v)
    s = str(v).strip()
    return s or None


def _parse_decimal(v, default: Decimal = Decimal("0")) -> Decimal:
    if v is None or v == "":
        return default
    try:
        return Decimal(str(v).strip())
    except (InvalidOperation, ValueError):
        return default


def _parse_part_type(v) -> models.PartType:
    s = _cell_str(v)
    if not s:
        return models.PartType.standard
    if s in PART_TYPE_IMPORT:
        return PART_TYPE_IMPORT[s]
    try:
        return models.PartType(s)
    except ValueError as e:
        raise ValueError(f"无法识别的类型：{s}（标准件/自制件/装配件 或 standard/custom/assembly）") from e


def _parse_status(v) -> models.StatusType:
    s = _cell_str(v)
    if not s:
        return models.StatusType.draft
    if s in STATUS_IMPORT:
        return STATUS_IMPORT[s]
    try:
        return models.StatusType(s)
    except ValueError as e:
        raise ValueError(f"无法识别的编辑状态：{s}（草稿/已发布 或 draft/released/obsolete）") from e


def _parse_bool_active(v) -> bool:
    if v is None or v == "":
        return True
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("是", "启用", "true", "1", "yes", "y"):
        return True
    if s in ("否", "停用", "false", "0", "no", "n"):
        return False
    return True


def _row_dict(headers: list, row_vals: tuple | list) -> dict[str, object | None]:
    col_map: dict[str, int] = {}
    for i, h in enumerate(headers):
        key = MATERIAL_HEADER_TO_FIELD.get(_norm_header(h))
        if key and key not in col_map:
            col_map[key] = i
    out: dict[str, object | None] = {}
    rv = list(row_vals) if row_vals else []
    for field, idx in col_map.items():
        out[field] = rv[idx] if idx < len(rv) else None
    return out


def _build_material_create(raw: dict[str, object | None]) -> schemas.MaterialCreate:
    name = _cell_str(raw.get("name"))
    if not name:
        raise ValueError("名称不能为空")

    code = _cell_str(raw.get("code"))

    spec = _cell_str(raw.get("spec"))
    drawing_no = _cell_str(raw.get("drawing_no"))
    comb = _cell_str(raw.get("spec_drawing"))
    if comb:
        if " / " in comb:
            left, right = comb.split(" / ", 1)
            spec = left.strip() or spec
            drawing_no = right.strip() or drawing_no
        else:
            spec = comb or spec

    return schemas.MaterialCreate(
        code=code,
        name=name,
        spec=spec,
        material_type=_cell_str(raw.get("material_type")),
        package_name=_cell_str(raw.get("package_name")),
        storage_location=_cell_str(raw.get("storage_location")),
        unit=_cell_str(raw.get("unit")),
        category=_cell_str(raw.get("category")),
        part_type=_parse_part_type(raw.get("part_type")) if raw.get("part_type") not in (None, "") else models.PartType.standard,
        usage=_cell_str(raw.get("usage")),
        material_name_attr=_cell_str(raw.get("material_name_attr")),
        standard_attr=_cell_str(raw.get("standard_attr")),
        grade_attr=_cell_str(raw.get("grade_attr")),
        default_supplier=_cell_str(raw.get("default_supplier")),
        tax_rate=_cell_str(raw.get("tax_rate")),
        purchase_link=_cell_str(raw.get("purchase_link")),
        current_revision=_cell_str(raw.get("current_revision")),
        status=_parse_status(raw.get("status")),
        is_active=_parse_bool_active(raw.get("is_active")),
        current_stock=_parse_decimal(raw.get("current_stock")),
        safety_stock=_parse_decimal(raw.get("safety_stock")),
        unit_price=_parse_decimal(raw.get("unit_price")),
        remark=_cell_str(raw.get("remark")),
        drawing_no=drawing_no,
    )


@router.get("/materials/template")
def download_material_import_template():
    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_title("物料导入")
    ws.append(MATERIAL_IMPORT_HEADERS)
    for c in range(1, len(MATERIAL_IMPORT_HEADERS) + 1):
        ws.cell(1, c).font = Font(bold=True)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename_ascii = "material_import_template.xlsx"
    filename_cn = "物料导入模板.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename_ascii}"'}
    headers["Content-Disposition"] += f"; filename*=UTF-8''{quote(filename_cn)}"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/materials", response_model=schemas.MaterialImportSummary)
async def import_materials(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 文件")

    content = await file.read(MAX_IMPORT_BYTES + 1)
    if len(content) > MAX_IMPORT_BYTES:
        raise HTTPException(status_code=413, detail="文件过大（上限 5MB）")

    try:
        wb = load_workbook(BytesIO(content), read_only=False, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法读取 Excel：{e}") from e

    ws = wb.active
    if ws.max_row is None or ws.max_row < 1:
        raise HTTPException(status_code=400, detail="空表格")
    if ws.max_row < 2:
        return schemas.MaterialImportSummary(created=0, failed=0, errors=[])

    max_col = ws.max_column or 1
    headers = [ws.cell(1, c).value for c in range(1, max_col + 1)]
    mapped = {_norm_header(h) for h in headers if h is not None}
    if "名称" not in mapped and "name" not in mapped:
        raise HTTPException(status_code=400, detail="第一行表头须包含「名称」列（建议下载模板）")

    created = 0
    failed = 0
    errors: list[schemas.MaterialImportErrorItem] = []

    for r in range(2, ws.max_row + 1):
        row_vals = tuple(ws.cell(r, c).value for c in range(1, max_col + 1))
        if not any(v not in (None, "") for v in row_vals):
            continue
        raw = _row_dict(headers, row_vals)
        if not _cell_str(raw.get("name")):
            failed += 1
            errors.append(schemas.MaterialImportErrorItem(row=r, message="名称不能为空"))
            continue
        try:
            payload = _build_material_create(raw)
            crud.create_material(db, payload)
            created += 1
        except ValidationError as e:
            failed += 1
            errors.append(schemas.MaterialImportErrorItem(row=r, message=str(e.errors())[:500]))
        except HTTPException as e:
            failed += 1
            detail = e.detail
            msg = detail if isinstance(detail, str) else str(detail)
            errors.append(schemas.MaterialImportErrorItem(row=r, message=msg[:500]))
        except ValueError as e:
            failed += 1
            errors.append(schemas.MaterialImportErrorItem(row=r, message=str(e)[:500]))
        except Exception as e:
            failed += 1
            errors.append(schemas.MaterialImportErrorItem(row=r, message=str(e)[:500]))

    return schemas.MaterialImportSummary(created=created, failed=failed, errors=errors[:200])
